import openpyxl
import os

sample_dir = r"D:\TRIO2026\docs\report_samples"

# Deep analysis of all 3 files - show ALL rows
for fname in sorted(os.listdir(sample_dir)):
    if not fname.endswith('.xlsx'): continue
    path = os.path.join(sample_dir, fname)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    
    print(f"\n{'='*80}")
    print(f"FILE: {fname}")
    
    for sname in wb.sheetnames:
        ws = wb[sname]
        print(f"Sheet: '{sname}' (rows={ws.max_row}, cols={ws.max_column})")
        print("-"*80)
        
        for i, row in enumerate(ws.iter_rows(max_row=ws.max_row, values_only=False)):
            cells = []
            for c in row:
                if c.value is not None:
                    cells.append(f"{c.column_letter}{c.row}={repr(c.value)}")
            if cells:
                print(f"  R{i+1:3d}: {' | '.join(cells)}")
    
    wb.close()
