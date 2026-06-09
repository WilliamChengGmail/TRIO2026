import openpyxl, os

path = r"D:\TRIO2026\docs\report_samples\20260105_132949.xlsx"
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
ws = wb['Sheet1']
print(f"rows={ws.max_row}, cols={ws.max_column}")
for i, row in enumerate(ws.iter_rows(max_row=ws.max_row, values_only=False)):
    cells = []
    for c in row:
        if c.value is not None and str(c.value).strip():
            cells.append(f"{c.column_letter}{c.row}={repr(c.value)}")
    if cells:
        print(f"  R{i+1:3d}: {' | '.join(cells)}")
wb.close()
