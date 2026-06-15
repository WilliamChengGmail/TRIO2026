"""
TRIO2026 測試資料種子工具
========================
從一代機 Excel 報告逆向工程，產生測試用 data.db。
供二代機報表引擎開發使用。

資料來源分類:
  [QR]     = QR Code 掃描獲得
  [MANUAL] = 操作員手動輸入
  [MACHINE]= 機器運行後產生
  [CALC]   = 軟體計算產生
  [SYSTEM] = 系統自動記錄

製作者: Office of William
"""

import sqlite3
import openpyxl
import json
import re
import uuid
from pathlib import Path
from datetime import datetime


# ─── 設定 ───────────────────────────────────────────
EXCEL_DIR = Path(r"//vmware-host/Shared Folders/[TRIO] 專案/機台產出的excel報告/trio_data")
OUTPUT_DB = Path(__file__).parent / "test_data.db"


# ─── DDL ────────────────────────────────────────────
DDL = """
-- TestRecord: 每次 Assay Run 一筆
CREATE TABLE IF NOT EXISTS TestRecord (
    Id                          INTEGER PRIMARY KEY AUTOINCREMENT,
    RunId                       TEXT    NOT NULL UNIQUE,     -- [SYSTEM] 時間戳生成

    -- 報告類型
    ReportType                  TEXT,                        -- [SYSTEM] IntelliPlex / Custom
    FlowName                    TEXT    NOT NULL DEFAULT '',  -- [QR] 流程名稱

    -- 操作員審計
    OperatorUserId              INTEGER,                     -- [SYSTEM] 操作員 ID
    OperatorUsername             TEXT,                        -- [SYSTEM] 操作員帳號快照
    OperatorDisplayName         TEXT,                        -- [SYSTEM] 操作員顯示名稱
    RoleLevel                   INTEGER,                     -- [SYSTEM] 角色等級

    -- 設備資訊
    DeviceSerialNo              TEXT,                        -- [SYSTEM] 設備序號
    SoftwareVersion             TEXT,                        -- [SYSTEM] 軟體版本
    InstallationUuid            TEXT,                        -- [SYSTEM] 設備 UUID

    -- 實驗參數
    ExperimentDate              TEXT,                        -- [MACHINE] yyyy/MM/dd
    FunctionModulesSelected     TEXT,                        -- [SYSTEM] Custom 模式的功能模組
    ExtractionProgram           TEXT,                        -- [QR] 萃取程式名稱
    ExtractionKitLotNo          TEXT,                        -- [QR] 萃取試劑盒批號
    ExtractionSampleVolume      TEXT,                        -- [QR] 萃取樣本體積
    ElutionVolume               TEXT,                        -- [QR] 洗脫體積
    PcrPlateId                  TEXT,                        -- [MANUAL] PCR 板 ID
    PcrTotalNucleicAcidInput    TEXT,                        -- [QR] PCR 核酸輸入量
    IntelliPlexKit1Name         TEXT,                        -- [QR] Kit 1 產品名稱
    IntelliPlexKit1LotNo        TEXT,                        -- [QR] Kit 1 批號
    IntelliPlexKit2Name         TEXT,                        -- [QR] Kit 2 產品名稱
    IntelliPlexKit2LotNo        TEXT,                        -- [QR] Kit 2 批號
    CustomPcrSetupJson          TEXT,                        -- [QR+MANUAL] JSON: Rxn1~4 設定
    ProductCode                 TEXT,                        -- [QR] 產品編碼

    -- 運行狀態
    SampleCount                 INTEGER,                     -- [SYSTEM] 樣本數
    StartTime                   TEXT    NOT NULL DEFAULT '',  -- [SYSTEM] ISO 8601
    EndTime                     TEXT,                        -- [SYSTEM] ISO 8601
    Status                      TEXT    NOT NULL DEFAULT 'Completed',
    CurrentPhase                TEXT,                        -- [SYSTEM] 當前階段
    ProgressPercent             INTEGER,                     -- [SYSTEM] 進度百分比
    CurrentStep                 INTEGER,                     -- [SYSTEM] 當前步驟
    TotalSteps                  INTEGER,                     -- [SYSTEM] 總步驟數
    ErrorCode                   TEXT,
    ErrorMessage                TEXT,
    Notes                       TEXT,

    -- 擴充欄位
    SampleBitmap                TEXT,                        -- [SYSTEM] 樣本啟用位圖
    ReagentCount                INTEGER,                     -- [QR] 試劑組數
    ReagentInfoJson             TEXT,                        -- [QR] 試劑 QR 解析 JSON
    OptSampleVolume             REAL,                        -- [QR] 光學檢測取樣體積
    FunctionFlags               TEXT,                        -- [SYSTEM] 功能旗標 JSON
    FlowDefinitionJson          TEXT,                        -- [SYSTEM] flow 定義快照

    -- 時間戳
    CreatedAt                   TEXT    NOT NULL,             -- [SYSTEM]
    UpdatedAt                   TEXT    NOT NULL              -- [SYSTEM]
);
CREATE INDEX IF NOT EXISTS IX_TestRecord_RunId ON TestRecord(RunId);
CREATE INDEX IF NOT EXISTS IX_TestRecord_OperatorUserId ON TestRecord(OperatorUserId);
CREATE INDEX IF NOT EXISTS IX_TestRecord_ExperimentDate ON TestRecord(ExperimentDate);

-- SampleResult: 每個樣本一筆
CREATE TABLE IF NOT EXISTS SampleResult (
    Id                      INTEGER PRIMARY KEY AUTOINCREMENT,
    TestRecordId            INTEGER NOT NULL REFERENCES TestRecord(Id) ON DELETE CASCADE,

    -- 樣本識別
    SamplePosition          INTEGER,                     -- [MACHINE] 孔位 1-24
    SampleType              TEXT,                        -- [SYSTEM] "NC"/"PC"/"Ctrl1"/"Ctrl2"/"Sample"
    SourcePosition          INTEGER,                     -- [MACHINE] 原始索引
    SampleBarcode           TEXT,                        -- [QR] 條碼
    SampleId                TEXT,                        -- [MANUAL] Sample ID
    ElutionTubeId           TEXT,                        -- [MANUAL] 洗脫管 ID

    -- 量測結果
    Concentration           REAL,                        -- [MACHINE] 濃度 ng/μL
    ConcentrationDisplay    TEXT,                        -- [CALC] 顯示文字
    UtilizedElutedVolume    REAL,                        -- [CALC] 使用體積 μL
    Volume                  REAL,                        -- [CALC] 通用體積

    -- PCR 孔位
    PcrWellKit1             TEXT,                        -- [CALC] Kit1/Rxn1 (如 A1)
    PcrWellKit2             TEXT,                        -- [CALC] Kit2/Rxn2
    PcrWellRxn3             TEXT,                        -- [CALC] Rxn3 (Custom)
    PcrWellRxn4             TEXT,                        -- [CALC] Rxn4 (Custom)

    -- 品質
    QualityFlag             TEXT,                        -- [CALC] Pass/Fail/Recheck
    RawDataJson             TEXT,                        -- [MACHINE] 光學原始數據 JSON

    CreatedAt               TEXT NOT NULL                 -- [SYSTEM]
);
CREATE INDEX IF NOT EXISTS IX_SampleResult_TestRecordId ON SampleResult(TestRecordId);

-- RawMeasurement: 原始量測數據 (1:1 with TestRecord)
CREATE TABLE IF NOT EXISTS RawMeasurement (
    Id                      INTEGER PRIMARY KEY AUTOINCREMENT,
    TestRecordId            INTEGER NOT NULL UNIQUE REFERENCES TestRecord(Id) ON DELETE CASCADE,
    S1AdValue               INTEGER,                     -- [MACHINE] S1 標準品 A/D
    S2AdValue               INTEGER,                     -- [MACHINE] S2 標準品 A/D
    S1Concentration         REAL,                        -- [MACHINE] S1 濃度
    S2Concentration         REAL,                        -- [MACHINE] S2 濃度
    RawAdValuesJson         TEXT,                        -- [MACHINE] arg0 原始值
    ConcentrationRawJson    TEXT,                        -- [MACHINE] arg2 原始值
    CalibrationCurveJson    TEXT,                        -- [MACHINE] 校正曲線
    CreatedAt               TEXT NOT NULL
);

-- ReportSnapshot: 報告快照
CREATE TABLE IF NOT EXISTS ReportSnapshot (
    Id                      INTEGER PRIMARY KEY AUTOINCREMENT,
    TestRecordId            INTEGER NOT NULL REFERENCES TestRecord(Id) ON DELETE CASCADE,
    ReportType              TEXT    NOT NULL,
    GeneratedAt             TEXT    NOT NULL,
    GeneratedByUserId       INTEGER,
    GeneratedByUsername     TEXT,
    ContentJson             TEXT,
    ExcelFilePath           TEXT,
    PdfFilePath             TEXT,
    FormatVersion           TEXT,
    ChecksumSha256          TEXT
);
CREATE INDEX IF NOT EXISTS IX_ReportSnapshot_TestRecordId ON ReportSnapshot(TestRecordId);
"""


def parse_na(val):
    """將 N/A 或空值轉為 None"""
    if val is None:
        return None
    s = str(val).strip()
    if s in ("N/A", "", "0 μL"):
        return None
    return s


def parse_float(val):
    """安全轉換為 float"""
    if val is None:
        return None
    s = str(val).strip()
    if s in ("N/A", ""):
        return None
    # 處理 "< 1.00" 或 "> 50.00"
    cleaned = re.sub(r"[<> ]", "", s)
    try:
        return float(cleaned)
    except ValueError:
        return None


def parse_int(val):
    """安全轉換為 int"""
    if val is None:
        return None
    s = str(val).strip()
    if s in ("N/A", ""):
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def parse_intelliplex(ws, filename_stem):
    """解析 IntelliPlex Report (mode=1)"""
    now = datetime.utcnow().isoformat() + "Z"
    run_id = filename_stem  # 如 20260317_135504

    # ── TestRecord ──
    record = {
        "RunId": run_id,
        "ReportType": "IntelliPlex",
        "FlowName": parse_na(ws.cell(4, 2).value) or "",
        "ExperimentDate": parse_na(ws.cell(3, 2).value),               # [MACHINE]
        "ExtractionProgram": parse_na(ws.cell(4, 2).value),            # [QR]
        "ExtractionKitLotNo": parse_na(ws.cell(5, 2).value),          # [QR]
        "ExtractionSampleVolume": parse_na(ws.cell(6, 2).value),      # [QR]
        "ElutionVolume": parse_na(ws.cell(7, 2).value),               # [QR]
        "PcrTotalNucleicAcidInput": parse_na(ws.cell(8, 2).value),    # [QR]
        "IntelliPlexKit1Name": parse_na(ws.cell(9, 2).value),         # [QR]
        "IntelliPlexKit1LotNo": parse_na(ws.cell(10, 2).value),      # [QR]
        "IntelliPlexKit2Name": parse_na(ws.cell(11, 2).value),        # [QR]
        "IntelliPlexKit2LotNo": parse_na(ws.cell(12, 2).value),      # [QR]
        "PcrPlateId": parse_na(ws.cell(13, 2).value),                 # [MANUAL]
        "StartTime": "",                                          # 逆向工程無法得知
        "EndTime": _run_id_to_iso(run_id),                            # RunId = 實驗完成時間
        "Status": "Completed",
        "CreatedAt": now,
        "UpdatedAt": now,
    }

    # ── RawMeasurement ──
    raw = {
        "S1AdValue": parse_int(ws.cell(14, 2).value),    # [MACHINE]
        "S2AdValue": parse_int(ws.cell(15, 2).value),     # [MACHINE]
        "CreatedAt": now,
    }

    # ── SampleResult ──
    samples = []
    for row in range(22, 46):  # Row 22~45
        a_val = ws.cell(row, 1).value
        if a_val is None or str(a_val).strip() == "":
            continue

        a_str = str(a_val).strip()

        # 判斷 SampleType
        if a_str in ("NC", "PC"):
            sample_type = a_str
            position = None
        else:
            sample_type = "Sample"
            position = parse_int(a_str)

        conc_raw = ws.cell(row, 2).value
        conc_display = parse_na(conc_raw)
        conc_val = parse_float(conc_raw)

        sample = {
            "SamplePosition": position,
            "SampleType": sample_type,
            "Concentration": conc_val,                          # [MACHINE]
            "ConcentrationDisplay": conc_display,               # [CALC]
            "UtilizedElutedVolume": parse_float(ws.cell(row, 3).value),  # [CALC]
            "PcrWellKit1": parse_na(ws.cell(row, 4).value),     # [CALC]
            "PcrWellKit2": parse_na(ws.cell(row, 5).value),     # [CALC]
            "SampleId": parse_na(ws.cell(row, 6).value),        # [MANUAL]
            "ElutionTubeId": parse_na(ws.cell(row, 7).value),   # [MANUAL]
            "CreatedAt": now,
        }
        samples.append(sample)

    record["SampleCount"] = len([s for s in samples if s["SampleType"] == "Sample"])
    return record, raw, samples


def parse_custom(ws, filename_stem):
    """解析 Custom Program Report (mode=2)"""
    now = datetime.utcnow().isoformat() + "Z"
    run_id = filename_stem

    # ── Custom PCR Setup (Row 10~15) → JSON ──
    pcr_setup = {}
    rxn_labels = ["Rxn1", "Rxn2", "Rxn3", "Rxn4"]
    for idx, label in enumerate(rxn_labels):
        col = 2 + idx  # B=2, C=3, D=4, E=5
        ctrl1 = parse_na(ws.cell(11, col).value)
        ctrl2 = parse_na(ws.cell(12, col).value)
        na_input = parse_na(ws.cell(13, col).value)
        smp_vol = parse_na(ws.cell(14, col).value)
        mm_vol = parse_na(ws.cell(15, col).value)

        if ctrl1 is None and na_input is None:
            pcr_setup[label] = None
        else:
            pcr_setup[label] = {
                "Ctrl1": ctrl1 == "Yes",
                "Ctrl2": ctrl2 == "Yes" if ctrl2 else False,
                "NucleicAcid": parse_float(na_input),
                "SampleVol": parse_float(smp_vol),
                "MasterMixVol": parse_float(mm_vol),
            }

    # ── TestRecord ──
    record = {
        "RunId": run_id,
        "ReportType": "Custom",
        "FlowName": "",
        "FunctionModulesSelected": parse_na(ws.cell(4, 2).value),      # [SYSTEM]
        "ExperimentDate": parse_na(ws.cell(3, 2).value),               # [MACHINE]
        "ExtractionProgram": parse_na(ws.cell(5, 2).value),            # [QR]
        "ExtractionKitLotNo": parse_na(ws.cell(6, 2).value),          # [QR]
        "ExtractionSampleVolume": parse_na(ws.cell(7, 2).value),      # [QR]
        "ElutionVolume": parse_na(ws.cell(8, 2).value),               # [QR]
        "PcrPlateId": parse_na(ws.cell(9, 2).value),                  # [MANUAL]
        "CustomPcrSetupJson": json.dumps(pcr_setup, ensure_ascii=False),
        "StartTime": "",                                          # 逆向工程無法得知
        "EndTime": _run_id_to_iso(run_id),                            # RunId = 實驗完成時間
        "Status": "Completed",
        "CreatedAt": now,
        "UpdatedAt": now,
    }

    # ── RawMeasurement ──
    raw = {
        "S1AdValue": parse_int(ws.cell(16, 2).value),
        "S2AdValue": parse_int(ws.cell(17, 2).value),
        "CreatedAt": now,
    }

    # ── SampleResult ──
    samples = []
    for row in range(24, 48):  # Row 24~47
        a_val = ws.cell(row, 1).value
        if a_val is None or str(a_val).strip() == "":
            continue

        a_str = str(a_val).strip()

        if a_str in ("Ctrl1", "Ctrl2"):
            sample_type = a_str
            position = None
        elif a_str in ("NC", "PC"):
            sample_type = a_str
            position = None
        else:
            sample_type = "Sample"
            position = parse_int(a_str)

        conc_raw = ws.cell(row, 2).value
        conc_display = parse_na(conc_raw)
        conc_val = parse_float(conc_raw)

        sample = {
            "SamplePosition": position,
            "SampleType": sample_type,
            "Concentration": conc_val,                              # [MACHINE]
            "ConcentrationDisplay": conc_display,                   # [CALC]
            "UtilizedElutedVolume": parse_float(ws.cell(row, 3).value),  # [CALC]
            "PcrWellKit1": parse_na(ws.cell(row, 4).value),         # [CALC] Rxn1
            "PcrWellKit2": parse_na(ws.cell(row, 5).value),         # [CALC] Rxn2
            "PcrWellRxn3": parse_na(ws.cell(row, 6).value),         # [CALC] Rxn3
            "PcrWellRxn4": parse_na(ws.cell(row, 7).value),         # [CALC] Rxn4
            "SampleId": parse_na(ws.cell(row, 8).value),            # [MANUAL]
            "ElutionTubeId": parse_na(ws.cell(row, 9).value),       # [MANUAL]
            "CreatedAt": now,
        }
        samples.append(sample)

    record["SampleCount"] = len([s for s in samples if s["SampleType"] == "Sample"])
    return record, raw, samples


def _run_id_to_iso(run_id):
    """20260317_135504 → 2026-03-17T13:55:04（此為實驗完成時間）"""
    try:
        # 去除可能的後綴（如 "20260120_155517 - 複製"）
        clean = run_id.split(" ")[0] if " " in run_id else run_id
        dt = datetime.strptime(clean, "%Y%m%d_%H%M%S")
        return dt.isoformat()
    except ValueError:
        return run_id


def create_db(db_path):
    """建立測試資料庫"""
    if db_path.exists():
        db_path.unlink()
    conn = sqlite3.connect(str(db_path))
    conn.executescript(DDL)
    return conn


def insert_record(conn, record):
    """插入 TestRecord，回傳 Id"""
    cols = [k for k in record.keys()]
    placeholders = ["?" for _ in cols]
    sql = f"INSERT INTO TestRecord ({','.join(cols)}) VALUES ({','.join(placeholders)})"
    cur = conn.execute(sql, [record[k] for k in cols])
    return cur.lastrowid


def insert_raw(conn, test_record_id, raw):
    """插入 RawMeasurement"""
    raw["TestRecordId"] = test_record_id
    cols = [k for k in raw.keys()]
    placeholders = ["?" for _ in cols]
    sql = f"INSERT INTO RawMeasurement ({','.join(cols)}) VALUES ({','.join(placeholders)})"
    conn.execute(sql, [raw[k] for k in cols])


def insert_samples(conn, test_record_id, samples):
    """插入 SampleResult"""
    for s in samples:
        s["TestRecordId"] = test_record_id
        cols = [k for k in s.keys()]
        placeholders = ["?" for _ in cols]
        sql = f"INSERT INTO SampleResult ({','.join(cols)}) VALUES ({','.join(placeholders)})"
        conn.execute(sql, [s[k] for k in cols])


def insert_snapshot(conn, test_record_id, record, samples):
    """插入 ReportSnapshot (將完整資料存為 JSON 快照)"""
    now = datetime.utcnow().isoformat() + "Z"
    content = {
        "header": {k: v for k, v in record.items()
                   if k not in ("CreatedAt", "UpdatedAt", "Status", "CurrentPhase",
                                "ProgressPercent", "CurrentStep", "TotalSteps")},
        "samples": samples,
    }
    conn.execute(
        """INSERT INTO ReportSnapshot
           (TestRecordId, ReportType, GeneratedAt, ContentJson, FormatVersion)
           VALUES (?, ?, ?, ?, ?)""",
        (test_record_id, record.get("ReportType", ""), now,
         json.dumps(content, ensure_ascii=False), "1.0-legacy")
    )


def _post_process(conn):
    """
    後處理：
    1. 為 38 筆 TestRecord 分配 OperatorUserId（operator:30 / admin:8）
    2. 模擬 2 筆 Error + 1 筆 Aborted 狀態
    """
    # ── 帳號定義（對應 UserSeed.cs） ──
    USERS = {
        "admin":    {"id": 1, "display": "Administrator", "role": 3},
        "operator": {"id": 3, "display": "Operator",      "role": 1},
    }

    # ── 分配策略：前 30 筆給 operator，後 8 筆給 admin ──
    all_ids = [row[0] for row in conn.execute(
        "SELECT Id FROM TestRecord ORDER BY Id"
    ).fetchall()]

    operator_ids = all_ids[:30]
    admin_ids = all_ids[30:]

    for tid in operator_ids:
        u = USERS["operator"]
        conn.execute(
            """UPDATE TestRecord
               SET OperatorUserId = ?, OperatorUsername = ?,
                   OperatorDisplayName = ?, RoleLevel = ?
               WHERE Id = ?""",
            (u["id"], "operator", u["display"], u["role"], tid)
        )

    for tid in admin_ids:
        u = USERS["admin"]
        conn.execute(
            """UPDATE TestRecord
               SET OperatorUserId = ?, OperatorUsername = ?,
                   OperatorDisplayName = ?, RoleLevel = ?
               WHERE Id = ?""",
            (u["id"], "admin", u["display"], u["role"], tid)
        )

    print(
        "  [POST] 分配操作員: operator={} 筆, admin={} 筆".format(
            len(operator_ids), len(admin_ids)
        )
    )

    # ── 模擬異常狀態（取 operator 的最後 3 筆） ──
    error_candidates = operator_ids[-3:]
    status_map = [
        (error_candidates[0], "Error",   "ERR-3001", "Optical read timeout"),
        (error_candidates[1], "Error",   "ERR-4002", "Pipette arm collision detected"),
        (error_candidates[2], "Aborted", None,       "User aborted during extraction"),
    ]

    for tid, status, code, msg in status_map:
        conn.execute(
            """UPDATE TestRecord
               SET Status = ?, ErrorCode = ?, ErrorMessage = ?
               WHERE Id = ?""",
            (status, code, msg, tid)
        )

    print("  [POST] 模擬異常狀態: 2 Error + 1 Aborted")


def main():
    print("=" * 60)
    print("TRIO2026 測試資料種子工具")
    print("從一代機 Excel 報告逆向工程產生 test_data.db")
    print("=" * 60)

    if not EXCEL_DIR.exists():
        print(f"[ERROR] Excel 目錄不存在: {EXCEL_DIR}")
        return

    xlsxs = sorted([f for f in EXCEL_DIR.glob("*.xlsx") if not f.name.startswith("~")])
    print(f"找到 {len(xlsxs)} 份 Excel 報告")

    conn = create_db(OUTPUT_DB)
    print(f"建立測試資料庫: {OUTPUT_DB}")

    success = 0
    errors = []

    for f in xlsxs:
        try:
            wb = openpyxl.load_workbook(str(f), read_only=True)
            ws = wb.active
            a1 = str(ws.cell(1, 1).value or "")

            stem = f.stem
            if "IntelliPlex" in a1:
                record, raw, samples = parse_intelliplex(ws, stem)
            else:
                record, raw, samples = parse_custom(ws, stem)

            wb.close()

            # 寫入 DB
            tid = insert_record(conn, record)
            insert_raw(conn, tid, raw)
            insert_samples(conn, tid, samples)
            insert_snapshot(conn, tid, record, samples)

            sample_count = len([s for s in samples if s.get("SampleType") == "Sample"])
            ctrl_count = len(samples) - sample_count
            print(f"  [OK] {stem} → {record['ReportType']:12s} "
                  f"| {sample_count} samples + {ctrl_count} ctrl")
            success += 1

        except Exception as e:
            errors.append((f.name, str(e)))
            print(f"  [ERR] {f.name}: {e}")

    # ══════════════════════════════════════
    # 後處理：分配操作員 + 模擬異常狀態
    # ══════════════════════════════════════
    _post_process(conn)

    conn.commit()

    # ── 統計 ──
    cur = conn.execute("SELECT COUNT(*) FROM TestRecord")
    tr_count = cur.fetchone()[0]
    cur = conn.execute("SELECT COUNT(*) FROM SampleResult")
    sr_count = cur.fetchone()[0]
    cur = conn.execute("SELECT COUNT(*) FROM RawMeasurement")
    rm_count = cur.fetchone()[0]
    cur = conn.execute("SELECT COUNT(*) FROM ReportSnapshot")
    rs_count = cur.fetchone()[0]

    conn.close()

    print()
    print("=" * 60)
    print(f"完成！成功: {success}, 失敗: {len(errors)}")
    print(f"  TestRecord:     {tr_count} 筆")
    print(f"  SampleResult:   {sr_count} 筆")
    print(f"  RawMeasurement: {rm_count} 筆")
    print(f"  ReportSnapshot: {rs_count} 筆")
    print(f"  資料庫: {OUTPUT_DB}")
    print("=" * 60)

    if errors:
        print("\n失敗清單:")
        for name, err in errors:
            print(f"  - {name}: {err}")


if __name__ == "__main__":
    main()
