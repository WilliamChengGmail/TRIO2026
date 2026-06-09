import openpyxl
import os
import json

sample_dir = r"D:\TRIO2026\docs\report_samples"
files = sorted([f for f in os.listdir(sample_dir) if f.endswith('.xlsx')])

for fname in files:
    path = os.path.join(sample_dir, fname)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    print(f"\n{'='*60}")
    print(f"FILE: {fname}")
    print(f"Sheets: {wb.sheetnames}")
    
    for sname in wb.sheetnames:
        ws = wb[sname]
        print(f"\n  --- Sheet: '{sname}' (rows={ws.max_row}, cols={ws.max_column}) ---")
        
        # Print first 5 rows to see headers and sample data
        for i, row in enumerate(ws.iter_rows(max_row=min(8, ws.max_row or 1), values_only=False)):
            cells = [(c.column_letter, c.value) for c in row if c.value is not None]
            if cells:
                print(f"  Row {i+1}: {cells}")
    
    wb.close()
